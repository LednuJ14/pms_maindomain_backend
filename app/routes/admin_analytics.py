"""
Admin Analytics Routes - Aggregates data from all property managers and properties
"""
import csv
import io
from datetime import datetime, timedelta
from flask import Blueprint, request, jsonify, current_app, make_response
from sqlalchemy import text, func
from app import db
from app.utils.decorators import admin_required
from app.utils.error_handlers import handle_api_error

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    current_app.logger.warning("reportlab not available. PDF reports will not work.")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    current_app.logger.warning("openpyxl not available. Excel reports will not work.")

admin_analytics_bp = Blueprint('admin_analytics', __name__)


def _range_start(range_key: str) -> datetime:
    """Get the start date for the given period."""
    now = datetime.utcnow()
    key = (range_key or '30days').lower()
    if key in ['7days', 'week']:
        return now - timedelta(days=7)
    if key in ['30days', 'month']:
        return now - timedelta(days=30)
    if key in ['90days', 'quarter']:
        return now - timedelta(days=90)
    if key in ['1year', 'year']:
        return now - timedelta(days=365)
    return now - timedelta(days=30)


@admin_analytics_bp.route('/analytics', methods=['GET'])
@admin_required
def get_admin_analytics(current_user):
    """
    Get admin analytics
    ---
    tags:
      - Admin Analytics
    summary: Get comprehensive analytics data (Admin only)
    description: Retrieve aggregated analytics data from all property managers and properties
    security:
      - Bearer: []
    parameters:
      - in: query
        name: property
        type: string
        description: Filter by property ID or 'all'
      - in: query
        name: range
        type: string
        description: Date range (30days, 7days, etc.)
    responses:
      200:
        description: Analytics data retrieved successfully
        schema:
          type: object
          properties:
            totalProperties:
              type: integer
            totalManagers:
              type: integer
            totalRevenue:
              type: number
      401:
        description: Unauthorized
      403:
        description: Forbidden - Admin access required
      500:
        description: Server error
    """
    try:
        property_filter = request.args.get('property', 'all')
        date_range = request.args.get('range', '30days')
        period_start = _range_start(date_range)
        
        # Build property filter
        property_where = ""
        property_params = {}
        
        if property_filter and str(property_filter).lower() != 'all':
            try:
                prop_id = int(property_filter)
                property_where = "WHERE p.id = :property_id"
                property_params['property_id'] = prop_id
            except (ValueError, TypeError):
                # Invalid property ID, ignore filter
                property_where = ""
                property_params = {}
        
        # Get all properties (or filtered)
        properties_sql = text(f"""
            SELECT p.id, p.title, p.building_name, p.status, p.owner_id,
                   u.first_name, u.last_name, u.email as owner_email
            FROM properties p
            LEFT JOIN users u ON p.owner_id = u.id
            {property_where}
            ORDER BY p.created_at DESC
        """)
        properties_result = db.session.execute(properties_sql, property_params).mappings().all()
        
        property_ids = [p['id'] for p in properties_result] if properties_result else []
        total_properties = len(property_ids)
        
        # If no properties, return safe defaults
        if not property_ids:
            # Still try to get manager count
            managers_sql = text(f"""
                SELECT COUNT(DISTINCT p.owner_id) as total_managers
                FROM properties p
                {property_where if property_where else ""}
            """)
            try:
                managers_result = db.session.execute(managers_sql, property_params).mappings().first()
                total_managers = int(managers_result['total_managers']) if managers_result else 0
            except Exception:
                total_managers = 0
            
            return jsonify({
                'totalProperties': 0,
                'totalRevenue': 0.0,
                'totalTenants': 0,
                'occupancyRate': 0.0,
                'maintenanceRequests': 0,
                'newInquiries': 0,
                'totalManagers': total_managers,
                'propertyPerformance': [],
                'managerPerformance': [],
                'monthlyData': []
            }), 200
        
        # Use tuple for IN clause (MySQL/MariaDB compatible)
        # Ensure it's always a tuple (single item needs trailing comma)
        property_ids_tuple = tuple(property_ids) if len(property_ids) > 1 else (property_ids[0],)
        
        # Calculate total revenue from tenant_units (active leases across all properties)
        revenue_sql = text("""
            SELECT COALESCE(SUM(tu.monthly_rent), 0) as total_revenue
            FROM tenant_units tu
            INNER JOIN units u ON u.id = tu.unit_id
            WHERE u.property_id IN :property_ids
            AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
        """)
        revenue_result = db.session.execute(revenue_sql, {'property_ids': property_ids_tuple}).mappings().first()
        total_revenue = float(revenue_result['total_revenue']) if revenue_result else 0.0
        
        # Calculate occupancy: count occupied units vs total units across all properties
        occupancy_sql = text("""
            SELECT 
                COUNT(DISTINCT u.id) as total_units,
                COUNT(DISTINCT CASE 
                    WHEN tu.id IS NOT NULL AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
                    THEN u.id 
                END) as occupied_units
            FROM units u
            LEFT JOIN tenant_units tu ON tu.unit_id = u.id 
                AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
            WHERE u.property_id IN :property_ids
        """)
        occupancy_result = db.session.execute(occupancy_sql, {'property_ids': property_ids_tuple}).mappings().first()
        total_units = int(occupancy_result['total_units']) if occupancy_result else 0
        occupied_units = int(occupancy_result['occupied_units']) if occupancy_result else 0
        occupancy_rate = round((occupied_units / total_units * 100), 2) if total_units > 0 else 0.0
        
        # Get total distinct tenants (active leases)
        tenants_sql = text("""
            SELECT COUNT(DISTINCT tu.tenant_id) as total_tenants
            FROM tenant_units tu
            INNER JOIN units u ON u.id = tu.unit_id
            WHERE u.property_id IN :property_ids
            AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
        """)
        tenants_result = db.session.execute(tenants_sql, {'property_ids': property_ids_tuple}).mappings().first()
        total_tenants = int(tenants_result['total_tenants']) if tenants_result else 0
        
        # Get maintenance requests count (within period)
        maintenance_requests = 0
        try:
            maintenance_sql = text("""
                SELECT COUNT(*) as total_requests
                FROM maintenance_requests mr
                WHERE mr.property_id IN :property_ids
                AND mr.created_at >= :period_start
            """)
            maintenance_result = db.session.execute(
                maintenance_sql, 
                {'property_ids': property_ids_tuple, 'period_start': period_start}
            ).mappings().first()
            maintenance_requests = int(maintenance_result['total_requests']) if maintenance_result else 0
        except Exception as e:
            current_app.logger.warning(f'Error fetching maintenance requests: {e}')
            maintenance_requests = 0
        
        # Get new inquiries within period
        new_inquiries = 0
        try:
            inquiries_sql = text("""
                SELECT COUNT(*) as total_inquiries
                FROM inquiries i
                WHERE i.property_id IN :property_ids
                AND i.created_at >= :period_start
            """)
            inquiries_result = db.session.execute(
                inquiries_sql,
                {'property_ids': property_ids_tuple, 'period_start': period_start}
            ).mappings().first()
            new_inquiries = int(inquiries_result['total_inquiries']) if inquiries_result else 0
        except Exception as e:
            current_app.logger.warning(f'Error fetching inquiries: {e}')
            new_inquiries = 0
        
        # Get total managers count
        managers_sql = text("""
            SELECT COUNT(DISTINCT p.owner_id) as total_managers
            FROM properties p
            {property_where}
        """.format(property_where=property_where if property_where else ""))
        managers_result = db.session.execute(managers_sql, property_params).mappings().first()
        total_managers = int(managers_result['total_managers']) if managers_result else 0
        
        # Get property performance breakdown
        property_performance = []
        for prop in properties_result[:20]:  # Limit to 20 properties to prevent performance issues
            try:
                prop_id = prop['id']
                
                # Get units count and occupied count for this property
                prop_units_sql = text("""
                    SELECT 
                        COUNT(DISTINCT u.id) as total_units,
                        COUNT(DISTINCT CASE 
                            WHEN tu.id IS NOT NULL AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
                            THEN u.id 
                        END) as occupied_units,
                        COALESCE(SUM(CASE 
                            WHEN tu.id IS NOT NULL AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
                            THEN tu.monthly_rent 
                            ELSE 0 
                        END), 0) as revenue
                    FROM units u
                    LEFT JOIN tenant_units tu ON tu.unit_id = u.id 
                        AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
                    WHERE u.property_id = :prop_id
                """)
                prop_units_result = db.session.execute(prop_units_sql, {'prop_id': prop_id}).mappings().first()
                
                prop_total_units = int(prop_units_result['total_units']) if prop_units_result else 0
                prop_occupied_units = int(prop_units_result['occupied_units']) if prop_units_result else 0
                prop_revenue = float(prop_units_result['revenue']) if prop_units_result else 0.0
                prop_occupancy = round((prop_occupied_units / prop_total_units * 100), 2) if prop_total_units > 0 else 0.0
                
                property_performance.append({
                    'id': prop_id,
                    'name': prop.get('title') or prop.get('building_name') or f'Property {prop_id}',
                    'occupancy': prop_occupancy,
                    'revenue': round(prop_revenue, 2),
                    'totalUnits': prop_total_units,
                    'occupiedUnits': prop_occupied_units,
                    'status': str(prop.get('status', '')).lower()
                })
            except Exception as e:
                current_app.logger.warning(f'Error processing property {prop.get("id")}: {e}')
                continue
        
        # Get manager performance breakdown
        manager_performance = []
        try:
            # Get distinct managers from properties
            distinct_managers_sql = text(f"""
                SELECT DISTINCT p.owner_id, u.first_name, u.last_name, u.email
                FROM properties p
                LEFT JOIN users u ON p.owner_id = u.id
                {property_where}
                WHERE p.owner_id IS NOT NULL
            """)
            managers_list = db.session.execute(distinct_managers_sql, property_params).mappings().all()
            
            for manager in managers_list[:20]:  # Limit to 20 managers
                try:
                    manager_id = manager['owner_id']
                    
                    # Get manager's properties
                    manager_props_sql = text("""
                        SELECT id FROM properties WHERE owner_id = :manager_id
                    """)
                    manager_props = db.session.execute(manager_props_sql, {'manager_id': manager_id}).fetchall()
                    manager_prop_ids = [p[0] for p in manager_props] if manager_props else []
                    
                    if not manager_prop_ids:
                        continue
                    
                    manager_prop_ids_tuple = tuple(manager_prop_ids)
                    
                    # Get manager's revenue
                    manager_revenue_sql = text("""
                        SELECT COALESCE(SUM(tu.monthly_rent), 0) as total_revenue
                        FROM tenant_units tu
                        INNER JOIN units u ON u.id = tu.unit_id
                        WHERE u.property_id IN :property_ids
                        AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
                    """)
                    manager_revenue_result = db.session.execute(
                        manager_revenue_sql, 
                        {'property_ids': manager_prop_ids_tuple}
                    ).mappings().first()
                    manager_revenue = float(manager_revenue_result['total_revenue']) if manager_revenue_result else 0.0
                    
                    manager_performance.append({
                        'id': manager_id,
                        'name': f"{manager.get('first_name', '')} {manager.get('last_name', '')}".strip() or manager.get('email', 'Unknown'),
                        'email': manager.get('email', ''),
                        'propertyCount': len(manager_prop_ids),
                        'revenue': round(manager_revenue, 2)
                    })
                except Exception as e:
                    current_app.logger.warning(f'Error processing manager {manager.get("owner_id")}: {e}')
                    continue
        except Exception as e:
            current_app.logger.warning(f'Error fetching manager performance: {e}')
            manager_performance = []
        
        # Generate monthly data (last 3 months)
        monthly_data = []
        try:
            for i in range(3):
                month_date = datetime.utcnow() - timedelta(days=30 * (2 - i))
                month_start = month_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                if i == 2:
                    month_end = datetime.utcnow()
                else:
                    next_month = month_date + timedelta(days=32)
                    month_end = next_month.replace(day=1) - timedelta(days=1)
                
                month_revenue_sql = text("""
                    SELECT COALESCE(SUM(tu.monthly_rent), 0) as revenue
                    FROM tenant_units tu
                    INNER JOIN units u ON u.id = tu.unit_id
                    WHERE u.property_id IN :property_ids
                    AND tu.move_in_date <= :month_end
                    AND (tu.move_out_date IS NULL OR tu.move_out_date >= :month_start)
                """)
                month_revenue_result = db.session.execute(
                    month_revenue_sql,
                    {'property_ids': property_ids_tuple, 'month_start': month_start, 'month_end': month_end}
                ).mappings().first()
                month_revenue = float(month_revenue_result['revenue']) if month_revenue_result else 0.0
                
                monthly_data.append({
                    'month': month_date.strftime('%b %Y'),
                    'revenue': round(month_revenue, 2)
                })
        except Exception as e:
            current_app.logger.warning(f'Error generating monthly data: {e}')
            monthly_data = []
        
        result = {
            'totalProperties': total_properties,
            'totalRevenue': round(total_revenue, 2),
            'totalTenants': total_tenants,
            'occupancyRate': occupancy_rate,
            'maintenanceRequests': maintenance_requests,
            'newInquiries': new_inquiries,
            'totalManagers': total_managers,
            'propertyPerformance': property_performance,
            'managerPerformance': manager_performance,
            'monthlyData': monthly_data
        }
        
        return jsonify(result), 200

    except Exception as e:
        current_app.logger.error(f'Get admin analytics error: {e}', exc_info=True)
        # Return safe defaults to prevent frontend crashes
        return jsonify({
            'totalProperties': 0,
            'totalRevenue': 0.0,
            'totalTenants': 0,
            'occupancyRate': 0.0,
            'maintenanceRequests': 0,
            'newInquiries': 0,
            'totalManagers': 0,
            'propertyPerformance': [],
            'managerPerformance': [],
            'monthlyData': [],
            'error': 'Failed to retrieve analytics data'
        }), 200  # Return 200 with error in response to prevent frontend crashes


def _get_admin_analytics_data(property_filter='all', date_range='30days'):
    """Helper function to get admin analytics data for reports."""
    period_start = _range_start(date_range)
    
    # Build property filter
    property_where = ""
    property_params = {}
    
    if property_filter and str(property_filter).lower() != 'all':
        try:
            prop_id = int(property_filter)
            property_where = "WHERE p.id = :property_id"
            property_params['property_id'] = prop_id
        except (ValueError, TypeError):
            property_where = ""
            property_params = {}
    
    # Get all properties (or filtered)
    properties_sql = text(f"""
        SELECT p.id, p.title, p.building_name, p.status, p.owner_id
        FROM properties p
        {property_where}
        ORDER BY p.created_at DESC
    """)
    properties_result = db.session.execute(properties_sql, property_params).mappings().all()
    
    property_ids = [p['id'] for p in properties_result] if properties_result else []
    
    if not property_ids:
        return {
            'totalProperties': 0,
            'totalRevenue': 0.0,
            'totalTenants': 0,
            'occupancyRate': 0.0,
            'maintenanceRequests': 0,
            'newInquiries': 0,
            'totalManagers': 0,
            'propertyPerformance': [],
            'managerPerformance': [],
            'monthlyData': [],
            'generated_at': datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
        }
    
    property_ids_tuple = tuple(property_ids) if len(property_ids) > 1 else (property_ids[0],)
    
    # Calculate total revenue
    revenue_sql = text("""
        SELECT COALESCE(SUM(tu.monthly_rent), 0) as total_revenue
        FROM tenant_units tu
        INNER JOIN units u ON u.id = tu.unit_id
        WHERE u.property_id IN :property_ids
        AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
    """)
    revenue_result = db.session.execute(revenue_sql, {'property_ids': property_ids_tuple}).mappings().first()
    total_revenue = float(revenue_result['total_revenue']) if revenue_result else 0.0
    
    # Calculate occupancy
    occupancy_sql = text("""
        SELECT 
            COUNT(DISTINCT u.id) as total_units,
            COUNT(DISTINCT CASE 
                WHEN tu.id IS NOT NULL AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
                THEN u.id 
            END) as occupied_units
        FROM units u
        LEFT JOIN tenant_units tu ON tu.unit_id = u.id 
            AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
        WHERE u.property_id IN :property_ids
    """)
    occupancy_result = db.session.execute(occupancy_sql, {'property_ids': property_ids_tuple}).mappings().first()
    total_units = int(occupancy_result['total_units']) if occupancy_result else 0
    occupied_units = int(occupancy_result['occupied_units']) if occupancy_result else 0
    occupancy_rate = round((occupied_units / total_units * 100), 2) if total_units > 0 else 0.0
    
    # Get total tenants
    tenants_sql = text("""
        SELECT COUNT(DISTINCT tu.tenant_id) as total_tenants
        FROM tenant_units tu
        INNER JOIN units u ON u.id = tu.unit_id
        WHERE u.property_id IN :property_ids
        AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
    """)
    tenants_result = db.session.execute(tenants_sql, {'property_ids': property_ids_tuple}).mappings().first()
    total_tenants = int(tenants_result['total_tenants']) if tenants_result else 0
    
    # Get maintenance requests
    maintenance_requests = 0
    try:
        maintenance_sql = text("""
            SELECT COUNT(*) as total_requests
            FROM maintenance_requests mr
            WHERE mr.property_id IN :property_ids
            AND mr.created_at >= :period_start
        """)
        maintenance_result = db.session.execute(
            maintenance_sql, 
            {'property_ids': property_ids_tuple, 'period_start': period_start}
        ).mappings().first()
        maintenance_requests = int(maintenance_result['total_requests']) if maintenance_result else 0
    except Exception:
        maintenance_requests = 0
    
    # Get new inquiries
    new_inquiries = 0
    try:
        inquiries_sql = text("""
            SELECT COUNT(*) as total_inquiries
            FROM inquiries i
            WHERE i.property_id IN :property_ids
            AND i.created_at >= :period_start
        """)
        inquiries_result = db.session.execute(
            inquiries_sql,
            {'property_ids': property_ids_tuple, 'period_start': period_start}
        ).mappings().first()
        new_inquiries = int(inquiries_result['total_inquiries']) if inquiries_result else 0
    except Exception:
        new_inquiries = 0
    
    # Get total managers
    managers_sql = text(f"""
        SELECT COUNT(DISTINCT p.owner_id) as total_managers
        FROM properties p
        {property_where if property_where else ""}
    """)
    managers_result = db.session.execute(managers_sql, property_params).mappings().first()
    total_managers = int(managers_result['total_managers']) if managers_result else 0
    
    # Get property performance
    property_performance = []
    for prop in properties_result[:20]:
        try:
            prop_id = prop['id']
            prop_units_sql = text("""
                SELECT 
                    COUNT(DISTINCT u.id) as total_units,
                    COUNT(DISTINCT CASE 
                        WHEN tu.id IS NOT NULL AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
                        THEN u.id 
                    END) as occupied_units,
                    COALESCE(SUM(CASE 
                        WHEN tu.id IS NOT NULL AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
                        THEN tu.monthly_rent 
                        ELSE 0 
                    END), 0) as revenue
                FROM units u
                LEFT JOIN tenant_units tu ON tu.unit_id = u.id 
                    AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
                WHERE u.property_id = :prop_id
            """)
            prop_units_result = db.session.execute(prop_units_sql, {'prop_id': prop_id}).mappings().first()
            
            prop_total_units = int(prop_units_result['total_units']) if prop_units_result else 0
            prop_occupied_units = int(prop_units_result['occupied_units']) if prop_units_result else 0
            prop_revenue = float(prop_units_result['revenue']) if prop_units_result else 0.0
            prop_occupancy = round((prop_occupied_units / prop_total_units * 100), 2) if prop_total_units > 0 else 0.0
            
            property_performance.append({
                'property': prop.get('title') or prop.get('building_name') or f'Property {prop_id}',
                'occupancy': prop_occupancy,
                'revenue': round(prop_revenue, 2)
            })
        except Exception:
            continue
    
    # Get manager performance
    manager_performance = []
    try:
        distinct_managers_sql = text(f"""
            SELECT DISTINCT p.owner_id, u.first_name, u.last_name, u.email
            FROM properties p
            LEFT JOIN users u ON p.owner_id = u.id
            {property_where}
            WHERE p.owner_id IS NOT NULL
        """)
        managers_list = db.session.execute(distinct_managers_sql, property_params).mappings().all()
        
        for manager in managers_list[:20]:
            try:
                manager_id = manager['owner_id']
                manager_props_sql = text("SELECT id FROM properties WHERE owner_id = :manager_id")
                manager_props = db.session.execute(manager_props_sql, {'manager_id': manager_id}).fetchall()
                manager_prop_ids = [p[0] for p in manager_props] if manager_props else []
                
                if not manager_prop_ids:
                    continue
                
                manager_prop_ids_tuple = tuple(manager_prop_ids)
                manager_revenue_sql = text("""
                    SELECT COALESCE(SUM(tu.monthly_rent), 0) as total_revenue
                    FROM tenant_units tu
                    INNER JOIN units u ON u.id = tu.unit_id
                    WHERE u.property_id IN :property_ids
                    AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
                """)
                manager_revenue_result = db.session.execute(
                    manager_revenue_sql, 
                    {'property_ids': manager_prop_ids_tuple}
                ).mappings().first()
                manager_revenue = float(manager_revenue_result['total_revenue']) if manager_revenue_result else 0.0
                
                manager_performance.append({
                    'name': f"{manager.get('first_name', '')} {manager.get('last_name', '')}".strip() or manager.get('email', 'Unknown'),
                    'email': manager.get('email', ''),
                    'propertyCount': len(manager_prop_ids),
                    'revenue': round(manager_revenue, 2)
                })
            except Exception:
                continue
    except Exception:
        manager_performance = []
    
    # Generate monthly data
    monthly_data = []
    try:
        for i in range(3):
            month_date = datetime.utcnow() - timedelta(days=30 * (2 - i))
            month_start = month_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            if i == 2:
                month_end = datetime.utcnow()
            else:
                next_month = month_date + timedelta(days=32)
                month_end = next_month.replace(day=1) - timedelta(days=1)
            
            month_revenue_sql = text("""
                SELECT COALESCE(SUM(tu.monthly_rent), 0) as revenue
                FROM tenant_units tu
                INNER JOIN units u ON u.id = tu.unit_id
                WHERE u.property_id IN :property_ids
                AND tu.move_in_date <= :month_end
                AND (tu.move_out_date IS NULL OR tu.move_out_date >= :month_start)
            """)
            month_revenue_result = db.session.execute(
                month_revenue_sql,
                {'property_ids': property_ids_tuple, 'month_start': month_start, 'month_end': month_end}
            ).mappings().first()
            month_revenue = float(month_revenue_result['revenue']) if month_revenue_result else 0.0
            
            monthly_data.append({
                'month': month_date.strftime('%b %Y'),
                'revenue': round(month_revenue, 2),
                'expenses': 0.0  # Admin doesn't track expenses per property manager
            })
    except Exception:
        monthly_data = []
    
    return {
        'totalProperties': len(property_ids),
        'totalRevenue': round(total_revenue, 2),
        'totalExpenses': 0.0,  # Admin doesn't track expenses
        'netIncome': round(total_revenue, 2),
        'totalTenants': total_tenants,
        'occupancyRate': occupancy_rate,
        'maintenanceRequests': maintenance_requests,
        'newInquiries': new_inquiries,
        'totalManagers': total_managers,
        'propertyPerformance': property_performance,
        'managerPerformance': manager_performance,
        'monthlyData': monthly_data,
        'generated_at': datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
    }


@admin_analytics_bp.route('/analytics/download/pdf', methods=['GET'])
@admin_required
def download_pdf_report(current_user):
    """Download admin analytics report as PDF."""
    if not REPORTLAB_AVAILABLE:
        return jsonify({'error': 'PDF generation not available. Please install reportlab.'}), 503
    
    try:
        from flask import send_file
        
        property_filter = request.args.get('property', 'all')
        date_range = request.args.get('range', '30days')
        report_type = request.args.get('report_type', 'overview')
        
        data = _get_admin_analytics_data(property_filter, date_range)
        
        # Create PDF in memory
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=1
        )
        story.append(Paragraph("Admin Analytics Report", title_style))
        
        # Report info
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#666666'),
            alignment=1
        )
        range_label = date_range.replace('_', ' ').title()
        property_label = 'All Properties' if property_filter == 'all' else f'Property ID: {property_filter}'
        story.append(Paragraph(f"Period: {range_label} | {property_label}", info_style))
        story.append(Paragraph(f"Generated: {data['generated_at']}", info_style))
        story.append(Spacer(1, 0.3*inch))
        
        # Key Metrics Table
        metrics_data = [
            ['Metric', 'Value'],
            ['Total Properties', str(data['totalProperties'])],
            ['Total Revenue', f"₱{data['totalRevenue']:,.2f}"],
            ['Total Tenants', str(data['totalTenants'])],
            ['Occupancy Rate', f"{data['occupancyRate']:.2f}%"],
            ['Maintenance Requests', str(data['maintenanceRequests'])],
            ['New Inquiries', str(data['newInquiries'])],
            ['Total Managers', str(data['totalManagers'])]
        ]
        
        metrics_table = Table(metrics_data, colWidths=[3*inch, 2*inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a1a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ]))
        story.append(Paragraph("Key Metrics", styles['Heading2']))
        story.append(Spacer(1, 0.1*inch))
        story.append(metrics_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Monthly Data
        if data['monthlyData']:
            story.append(Paragraph("Monthly Revenue", styles['Heading2']))
            story.append(Spacer(1, 0.1*inch))
            monthly_data_table = [['Month', 'Revenue']]
            for month in data['monthlyData']:
                monthly_data_table.append([
                    month['month'],
                    f"₱{month['revenue']:,.2f}"
                ])
            
            monthly_table = Table(monthly_data_table, colWidths=[3*inch, 2*inch])
            monthly_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a1a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
            ]))
            story.append(monthly_table)
            story.append(Spacer(1, 0.3*inch))
        
        # Property Performance
        if data['propertyPerformance']:
            story.append(Paragraph("Property Performance", styles['Heading2']))
            story.append(Spacer(1, 0.1*inch))
            prop_data_table = [['Property', 'Occupancy %', 'Monthly Revenue']]
            for prop in data['propertyPerformance']:
                prop_data_table.append([
                    prop['property'],
                    f"{prop['occupancy']:.2f}%",
                    f"₱{prop['revenue']:,.2f}"
                ])
            
            prop_table = Table(prop_data_table, colWidths=[3*inch, 1.5*inch, 1.5*inch])
            prop_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a1a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
            ]))
            story.append(prop_table)
            story.append(Spacer(1, 0.3*inch))
        
        # Manager Performance
        if data['managerPerformance']:
            story.append(PageBreak())
            story.append(Paragraph("Manager Performance", styles['Heading2']))
            story.append(Spacer(1, 0.1*inch))
            manager_data_table = [['Manager Name', 'Email', 'Properties', 'Total Revenue']]
            for manager in data['managerPerformance']:
                manager_data_table.append([
                    manager['name'],
                    manager.get('email', ''),
                    str(manager['propertyCount']),
                    f"₱{manager['revenue']:,.2f}"
                ])
            
            manager_table = Table(manager_data_table, colWidths=[2.5*inch, 2*inch, 1*inch, 1.5*inch])
            manager_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a1a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
            ]))
            story.append(manager_table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        # Generate filename
        range_str = date_range.replace(' ', '_')
        filename = f"admin_analytics_report_{range_str}_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        current_app.logger.error(f'Admin PDF report generation error: {e}', exc_info=True)
        return handle_api_error(500, 'Failed to generate PDF report')


@admin_analytics_bp.route('/analytics/download/excel', methods=['GET'])
@admin_required
def download_excel_report(current_user):
    """Download admin analytics report as Excel."""
    if not OPENPYXL_AVAILABLE:
        return jsonify({'error': 'Excel generation not available. Please install openpyxl.'}), 503
    
    try:
        property_filter = request.args.get('property', 'all')
        date_range = request.args.get('range', '30days')
        
        data = _get_admin_analytics_data(property_filter, date_range)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Admin Analytics Report"
        
        # Styles
        header_fill = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=16)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = "Admin Analytics Report"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align
        
        # Report info
        range_label = date_range.replace('_', ' ').title()
        property_label = 'All Properties' if property_filter == 'all' else f'Property ID: {property_filter}'
        ws['A2'] = f"Period: {range_label} | {property_label}"
        ws['A3'] = f"Generated: {data['generated_at']}"
        
        row = 5
        
        # Key Metrics
        ws[f'A{row}'] = "Key Metrics"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        metrics_headers = ['Metric', 'Value']
        for col, header in enumerate(metrics_headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align
        
        row += 1
        metrics_data = [
            ['Total Properties', data['totalProperties']],
            ['Total Revenue', f"₱{data['totalRevenue']:,.2f}"],
            ['Total Tenants', data['totalTenants']],
            ['Occupancy Rate', f"{data['occupancyRate']:.2f}%"],
            ['Maintenance Requests', data['maintenanceRequests']],
            ['New Inquiries', data['newInquiries']],
            ['Total Managers', data['totalManagers']]
        ]
        
        for metric_row in metrics_data:
            for col, value in enumerate(metric_row, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = border
            row += 1
        
        row += 2
        
        # Monthly Data
        if data['monthlyData']:
            ws[f'A{row}'] = "Monthly Revenue"
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 1
            
            monthly_headers = ['Month', 'Revenue']
            for col, header in enumerate(monthly_headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center_align
            
            row += 1
            for month in data['monthlyData']:
                ws.cell(row=row, column=1).value = month['month']
                ws.cell(row=row, column=2).value = f"₱{month['revenue']:,.2f}"
                for col in range(1, 3):
                    ws.cell(row=row, column=col).border = border
                row += 1
            
            row += 2
        
        # Property Performance
        if data['propertyPerformance']:
            ws[f'A{row}'] = "Property Performance"
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 1
            
            prop_headers = ['Property', 'Occupancy %', 'Monthly Revenue']
            for col, header in enumerate(prop_headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center_align
            
            row += 1
            for prop in data['propertyPerformance']:
                ws.cell(row=row, column=1).value = prop['property']
                ws.cell(row=row, column=2).value = f"{prop['occupancy']:.2f}%"
                ws.cell(row=row, column=3).value = f"₱{prop['revenue']:,.2f}"
                for col in range(1, 4):
                    ws.cell(row=row, column=col).border = border
                row += 1
            
            row += 2
        
        # Manager Performance
        if data['managerPerformance']:
            ws[f'A{row}'] = "Manager Performance"
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 1
            
            manager_headers = ['Manager Name', 'Email', 'Properties', 'Total Revenue']
            for col, header in enumerate(manager_headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center_align
            
            row += 1
            for manager in data['managerPerformance']:
                ws.cell(row=row, column=1).value = manager['name']
                ws.cell(row=row, column=2).value = manager.get('email', '')
                ws.cell(row=row, column=3).value = manager['propertyCount']
                ws.cell(row=row, column=4).value = f"₱{manager['revenue']:,.2f}"
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = border
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Generate filename
        range_str = date_range.replace(' ', '_')
        filename = f"admin_analytics_report_{range_str}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
        
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        current_app.logger.error(f'Admin Excel report generation error: {e}', exc_info=True)
        return handle_api_error(500, 'Failed to generate Excel report')


@admin_analytics_bp.route('/analytics/download/csv', methods=['GET'])
@admin_required
def download_csv_report(current_user):
    """Download admin analytics report as CSV."""
    try:
        property_filter = request.args.get('property', 'all')
        date_range = request.args.get('range', '30days')
        
        data = _get_admin_analytics_data(property_filter, date_range)
        
        # Create CSV in memory
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        
        # Header
        writer.writerow(["Admin Analytics Report"])
        range_label = date_range.replace('_', ' ').title()
        property_label = 'All Properties' if property_filter == 'all' else f'Property ID: {property_filter}'
        writer.writerow([f"Period: {range_label} | {property_label}"])
        writer.writerow([f"Generated: {data['generated_at']}"])
        writer.writerow([])
        
        # Key Metrics
        writer.writerow(["Key Metrics"])
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Total Properties", str(data['totalProperties'])])
        writer.writerow(["Total Revenue", f"₱{data['totalRevenue']:,.2f}"])
        writer.writerow(["Total Tenants", str(data['totalTenants'])])
        writer.writerow(["Occupancy Rate", f"{data['occupancyRate']:.2f}%"])
        writer.writerow(["Maintenance Requests", str(data['maintenanceRequests'])])
        writer.writerow(["New Inquiries", str(data['newInquiries'])])
        writer.writerow(["Total Managers", str(data['totalManagers'])])
        writer.writerow([])
        
        # Monthly Data
        if data['monthlyData']:
            writer.writerow(["Monthly Revenue"])
            writer.writerow(["Month", "Revenue"])
            for month in data['monthlyData']:
                writer.writerow([
                    month['month'],
                    f"₱{month['revenue']:,.2f}"
                ])
            writer.writerow([])
        
        # Property Performance
        if data['propertyPerformance']:
            writer.writerow(["Property Performance"])
            writer.writerow(["Property", "Occupancy %", "Monthly Revenue"])
            for prop in data['propertyPerformance']:
                writer.writerow([
                    prop['property'],
                    f"{prop['occupancy']:.2f}%",
                    f"₱{prop['revenue']:,.2f}"
                ])
            writer.writerow([])
        
        # Manager Performance
        if data['managerPerformance']:
            writer.writerow(["Manager Performance"])
            writer.writerow(["Manager Name", "Email", "Properties", "Total Revenue"])
            for manager in data['managerPerformance']:
                writer.writerow([
                    manager['name'],
                    manager.get('email', ''),
                    str(manager['propertyCount']),
                    f"₱{manager['revenue']:,.2f}"
                ])
        
        # Convert to bytes
        csv_bytes = buffer.getvalue().encode('utf-8-sig')
        buffer.close()
        
        # Create response
        response = make_response(csv_bytes)
        range_str = date_range.replace(' ', '_')
        filename = f"admin_analytics_report_{range_str}_{datetime.utcnow().strftime('%Y%m%d')}.csv"
        response.headers['Content-Type'] = 'text/csv; charset=utf-8-sig'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        current_app.logger.error(f'Admin CSV report generation error: {e}', exc_info=True)
        return handle_api_error(500, 'Failed to generate CSV report')


