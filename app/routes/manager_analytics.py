"""
Manager Analytics API Routes
"""
import csv
import io
from datetime import datetime, timedelta
from flask import Blueprint, request, jsonify, current_app, send_file, make_response
from sqlalchemy import text, func
from app import db
from app.utils.decorators import manager_required
from app.utils.error_handlers import handle_api_error

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    current_app.logger.warning("reportlab not available. PDF reports will not work.")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    current_app.logger.warning("openpyxl not available. Excel reports will not work.")

manager_analytics_bp = Blueprint('manager_analytics', __name__)


def _get_period_start(period: str) -> datetime:
    """Get the start date for the given period."""
    now = datetime.utcnow()
    period = (period or 'month').lower()
    if period in ['week', '7days']:
        return now - timedelta(days=7)
    if period in ['month', '30days']:
        return now - timedelta(days=30)
    if period in ['quarter', '90days']:
        return now - timedelta(days=90)
    if period in ['year', '1year']:
        return now - timedelta(days=365)
    return now - timedelta(days=30)


@manager_analytics_bp.route('', methods=['GET'])
@manager_analytics_bp.route('/', methods=['GET'])
@manager_required
def get_manager_analytics(current_user):
    """
    Get manager analytics
    ---
    tags:
      - Manager Analytics
    summary: Get analytics data for the property manager
    description: Retrieve analytics data for properties owned by the authenticated property manager
    security:
      - Bearer: []
    parameters:
      - in: query
        name: property
        type: string
        description: Filter by property ID or 'all'
      - in: query
        name: period
        type: string
        description: Time period (month, week, year)
    responses:
      200:
        description: Analytics data retrieved successfully
        schema:
          type: object
          properties:
            totalRevenue:
              type: number
            totalExpenses:
              type: number
            netIncome:
              type: number
            occupancyRate:
              type: number
      401:
        description: Unauthorized
      403:
        description: Forbidden - Manager access required
      500:
        description: Server error
    """
    try:
        # Query params
        property_filter = request.args.get('property', 'all')
        period = request.args.get('period', 'month')
        period_start = _get_period_start(period)
        
        # Build property filter SQL
        property_where = "WHERE p.owner_id = :owner_id"
        property_params = {'owner_id': current_user.id}
        
        if property_filter and property_filter != 'all':
            try:
                property_id = int(property_filter)
                property_where += " AND p.id = :property_id"
                property_params['property_id'] = property_id
            except (ValueError, TypeError):
                # Invalid property ID, ignore filter
                pass
        
        # Get manager's properties
        properties_sql = text(f"""
            SELECT p.id, p.title, p.building_name, p.status
            FROM properties p
            {property_where}
        """)
        properties_result = db.session.execute(properties_sql, property_params).mappings().all()
        
        property_ids = [p['id'] for p in properties_result] if properties_result else []
        
        if not property_ids:
            # No properties, return empty analytics
            return jsonify({
                'totalRevenue': 0.0,
                'totalExpenses': 0.0,
                'netIncome': 0.0,
                'occupancyRate': 0.0,
                'maintenanceRequests': 0,
                'tenantSatisfaction': 0.0,
                'monthlyData': [],
                'propertyPerformance': []
            }), 200
        
        # Use tuple for IN clause (MySQL/MariaDB compatible)
        property_ids_tuple = tuple(property_ids)
        
        # Calculate real revenue from tenant_units (active leases)
        # Revenue represents total monthly rent from all currently active leases
        # This is the recurring monthly revenue, not filtered by period
        revenue_sql = text("""
            SELECT COALESCE(SUM(tu.monthly_rent), 0) as total_revenue
            FROM tenant_units tu
            INNER JOIN units u ON u.id = tu.unit_id
            WHERE u.property_id IN :property_ids
            AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
            AND tu.move_in_date <= NOW()
        """)
        revenue_result = db.session.execute(revenue_sql, {'property_ids': property_ids_tuple}).mappings().first()
        total_revenue = float(revenue_result['total_revenue']) if revenue_result else 0.0
        
        # Calculate occupancy: count occupied units vs total units
        occupancy_sql = text("""
            SELECT 
                COUNT(DISTINCT u.id) as total_units,
                COUNT(DISTINCT CASE 
                    WHEN tu.id IS NOT NULL AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
                    THEN u.id 
                END) as occupied_units
            FROM units u
            LEFT JOIN tenant_units tu ON tu.unit_id = u.id 
                AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
            WHERE u.property_id IN :property_ids
        """)
        occupancy_result = db.session.execute(occupancy_sql, {'property_ids': property_ids_tuple}).mappings().first()
        total_units = int(occupancy_result['total_units']) if occupancy_result else 0
        occupied_units = int(occupancy_result['occupied_units']) if occupancy_result else 0
        occupancy_rate = round((occupied_units / total_units * 100), 2) if total_units > 0 else 0.0
        
        # Get real maintenance requests count (within period)
        maintenance_sql = text("""
            SELECT COUNT(*) as total_requests
            FROM maintenance_requests mr
            WHERE mr.property_id IN :property_ids
            AND mr.created_at >= :period_start
        """)
        maintenance_result = db.session.execute(
            maintenance_sql, 
            {'property_ids': property_ids_tuple, 'period_start': period_start}
        ).mappings().first()
        maintenance_requests = int(maintenance_result['total_requests']) if maintenance_result else 0
        
        # Get tenant satisfaction from maintenance_requests ratings
        satisfaction_sql = text("""
            SELECT 
                AVG(mr.tenant_satisfaction_rating) as avg_rating,
                COUNT(mr.tenant_satisfaction_rating) as rating_count
            FROM maintenance_requests mr
            WHERE mr.property_id IN :property_ids
            AND mr.tenant_satisfaction_rating IS NOT NULL
            AND mr.created_at >= :period_start
        """)
        satisfaction_result = db.session.execute(
            satisfaction_sql,
            {'property_ids': property_ids_tuple, 'period_start': period_start}
        ).mappings().first()
        
        tenant_satisfaction = 0.0
        if satisfaction_result and satisfaction_result['avg_rating']:
            tenant_satisfaction = round(float(satisfaction_result['avg_rating']), 1)
        
        # Calculate real expenses from subscription bills (paid/verified bills within period)
        expenses_sql = text("""
            SELECT COALESCE(SUM(sb.amount), 0) as total_expenses
            FROM subscription_bills sb
            WHERE sb.user_id = :user_id
            AND sb.status IN ('paid', 'verified', 'completed')
            AND (
                (sb.billing_period_start >= :period_start AND sb.billing_period_start <= NOW())
                OR (sb.payment_date >= :period_start AND sb.payment_date <= NOW())
                OR (sb.created_at >= :period_start AND sb.created_at <= NOW())
            )
        """)
        expenses_result = db.session.execute(
            expenses_sql,
            {'user_id': current_user.id, 'period_start': period_start}
        ).mappings().first()
        total_expenses = float(expenses_result['total_expenses']) if expenses_result else 0.0
        total_expenses = round(total_expenses, 2)
        net_income = round(total_revenue - total_expenses, 2)
        
        # Get property performance data
        property_performance = []
        for prop in properties_result:
            prop_id = prop['id']
            
            # Get units count and occupied count for this property
            prop_units_sql = text("""
                SELECT 
                    COUNT(DISTINCT u.id) as total_units,
                    COUNT(DISTINCT CASE 
                        WHEN tu.id IS NOT NULL AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
                        THEN u.id 
                    END) as occupied_units,
                    COALESCE(SUM(CASE 
                        WHEN tu.id IS NOT NULL AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
                        THEN tu.monthly_rent 
                        ELSE 0 
                    END), 0) as revenue
                FROM units u
                LEFT JOIN tenant_units tu ON tu.unit_id = u.id 
                    AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
                WHERE u.property_id = :prop_id
            """)
            prop_units_result = db.session.execute(prop_units_sql, {'prop_id': prop_id}).mappings().first()
            
            prop_total_units = int(prop_units_result['total_units']) if prop_units_result else 0
            prop_occupied_units = int(prop_units_result['occupied_units']) if prop_units_result else 0
            prop_revenue = float(prop_units_result['revenue']) if prop_units_result else 0.0
            prop_occupancy = round((prop_occupied_units / prop_total_units * 100), 2) if prop_total_units > 0 else 0.0
            
            property_performance.append({
                'property': prop.get('title') or prop.get('building_name') or f'Property {prop_id}',
                'occupancy': prop_occupancy,
                'revenue': prop_revenue
            })
        
        # Generate monthly data (last 3 months)
        monthly_data = []
        for i in range(3):
            month_date = datetime.utcnow() - timedelta(days=30 * (2 - i))
            month_start = month_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            if i == 2:
                month_end = datetime.utcnow()
            else:
                next_month = month_date + timedelta(days=32)
                month_end = next_month.replace(day=1) - timedelta(days=1)
            
            month_revenue_sql = text("""
                SELECT COALESCE(SUM(tu.monthly_rent), 0) as revenue
                FROM tenant_units tu
                INNER JOIN units u ON u.id = tu.unit_id
                WHERE u.property_id IN :property_ids
                AND tu.move_in_date <= :month_end
                AND (tu.move_out_date IS NULL OR tu.move_out_date >= :month_start)
            """)
            month_revenue_result = db.session.execute(
                month_revenue_sql,
                {'property_ids': property_ids_tuple, 'month_start': month_start, 'month_end': month_end}
            ).mappings().first()
            month_revenue = float(month_revenue_result['revenue']) if month_revenue_result else 0.0
            
            # Calculate monthly expenses from subscription bills
            month_expenses_sql = text("""
                SELECT COALESCE(SUM(sb.amount), 0) as expenses
                FROM subscription_bills sb
                WHERE sb.user_id = :user_id
                AND sb.status IN ('paid', 'verified', 'completed')
                AND (
                    (sb.billing_period_start >= :month_start AND sb.billing_period_start <= :month_end)
                    OR (sb.payment_date >= :month_start AND sb.payment_date <= :month_end)
                    OR (sb.created_at >= :month_start AND sb.created_at <= :month_end)
                )
            """)
            month_expenses_result = db.session.execute(
                month_expenses_sql,
                {'user_id': current_user.id, 'month_start': month_start, 'month_end': month_end}
            ).mappings().first()
            month_expenses = float(month_expenses_result['expenses']) if month_expenses_result else 0.0
            month_expenses = round(month_expenses, 2)
            
            monthly_data.append({
                'month': month_date.strftime('%b'),
                'revenue': round(month_revenue, 2),
                'expenses': month_expenses
            })
        
        result = {
            'totalRevenue': round(total_revenue, 2),
            'totalExpenses': round(total_expenses, 2),
            'netIncome': round(net_income, 2),
            'occupancyRate': occupancy_rate,
            'maintenanceRequests': maintenance_requests,
            'tenantSatisfaction': tenant_satisfaction,
            'monthlyData': monthly_data,
            'propertyPerformance': property_performance[:10]  # Limit to 10 properties
        }

        return jsonify(result), 200

    except Exception as e:
        current_app.logger.error(f'Get manager analytics error: {e}', exc_info=True)
        # Return safe default values instead of crashing
        return jsonify({
            'totalRevenue': 0.0,
            'totalExpenses': 0.0,
            'netIncome': 0.0,
            'occupancyRate': 0.0,
            'maintenanceRequests': 0,
            'tenantSatisfaction': 0.0,
            'monthlyData': [],
            'propertyPerformance': [],
            'error': 'Failed to retrieve analytics data'
        }), 200  # Return 200 with error in response to prevent frontend crashes


def _get_analytics_data(current_user, property_filter='all', period='month'):
    """Helper function to get analytics data for report generation."""
    period_start = _get_period_start(period)
    
    # Build property filter SQL
    property_where = "WHERE p.owner_id = :owner_id"
    property_params = {'owner_id': current_user.id}
    
    if property_filter and property_filter != 'all':
        try:
            property_id = int(property_filter)
            property_where += " AND p.id = :property_id"
            property_params['property_id'] = property_id
        except (ValueError, TypeError):
            pass
    
    # Get manager's properties
    properties_sql = text(f"""
        SELECT p.id, p.title, p.building_name, p.status
        FROM properties p
        {property_where}
    """)
    properties_result = db.session.execute(properties_sql, property_params).mappings().all()
    
    property_ids = [p['id'] for p in properties_result] if properties_result else []
    
    if not property_ids:
        return {
            'properties': [],
            'totalRevenue': 0.0,
            'totalExpenses': 0.0,
            'netIncome': 0.0,
            'occupancyRate': 0.0,
            'maintenanceRequests': 0,
            'tenantSatisfaction': 0.0,
            'monthlyData': [],
            'propertyPerformance': [],
            'period': period,
            'property_filter': property_filter
        }
    
    property_ids_tuple = tuple(property_ids)
    
    # Calculate revenue
    revenue_sql = text("""
        SELECT COALESCE(SUM(tu.monthly_rent), 0) as total_revenue
        FROM tenant_units tu
        INNER JOIN units u ON u.id = tu.unit_id
        WHERE u.property_id IN :property_ids
        AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
        AND tu.move_in_date <= NOW()
    """)
    revenue_result = db.session.execute(revenue_sql, {'property_ids': property_ids_tuple}).mappings().first()
    total_revenue = float(revenue_result['total_revenue']) if revenue_result else 0.0
    
    # Calculate occupancy
    occupancy_sql = text("""
        SELECT 
            COUNT(DISTINCT u.id) as total_units,
            COUNT(DISTINCT CASE 
                WHEN tu.id IS NOT NULL AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
                THEN u.id 
            END) as occupied_units
        FROM units u
        LEFT JOIN tenant_units tu ON tu.unit_id = u.id 
            AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
        WHERE u.property_id IN :property_ids
    """)
    occupancy_result = db.session.execute(occupancy_sql, {'property_ids': property_ids_tuple}).mappings().first()
    total_units = int(occupancy_result['total_units']) if occupancy_result else 0
    occupied_units = int(occupancy_result['occupied_units']) if occupancy_result else 0
    occupancy_rate = round((occupied_units / total_units * 100), 2) if total_units > 0 else 0.0
    
    # Maintenance requests
    maintenance_sql = text("""
        SELECT COUNT(*) as total_requests
        FROM maintenance_requests mr
        WHERE mr.property_id IN :property_ids
        AND mr.created_at >= :period_start
    """)
    maintenance_result = db.session.execute(
        maintenance_sql, 
        {'property_ids': property_ids_tuple, 'period_start': period_start}
    ).mappings().first()
    maintenance_requests = int(maintenance_result['total_requests']) if maintenance_result else 0
    
    # Tenant satisfaction
    satisfaction_sql = text("""
        SELECT 
            AVG(mr.tenant_satisfaction_rating) as avg_rating,
            COUNT(mr.tenant_satisfaction_rating) as rating_count
        FROM maintenance_requests mr
        WHERE mr.property_id IN :property_ids
        AND mr.tenant_satisfaction_rating IS NOT NULL
        AND mr.created_at >= :period_start
    """)
    satisfaction_result = db.session.execute(
        satisfaction_sql,
        {'property_ids': property_ids_tuple, 'period_start': period_start}
    ).mappings().first()
    
    tenant_satisfaction = 0.0
    if satisfaction_result and satisfaction_result['avg_rating']:
        tenant_satisfaction = round(float(satisfaction_result['avg_rating']), 1)
    
    # Expenses
    expenses_sql = text("""
        SELECT COALESCE(SUM(sb.amount), 0) as total_expenses
        FROM subscription_bills sb
        WHERE sb.user_id = :user_id
        AND sb.status IN ('paid', 'verified', 'completed')
        AND (
            (sb.billing_period_start >= :period_start AND sb.billing_period_start <= NOW())
            OR (sb.payment_date >= :period_start AND sb.payment_date <= NOW())
            OR (sb.created_at >= :period_start AND sb.created_at <= NOW())
        )
    """)
    expenses_result = db.session.execute(
        expenses_sql,
        {'user_id': current_user.id, 'period_start': period_start}
    ).mappings().first()
    total_expenses = float(expenses_result['total_expenses']) if expenses_result else 0.0
    total_expenses = round(total_expenses, 2)
    net_income = round(total_revenue - total_expenses, 2)
    
    # Property performance
    property_performance = []
    for prop in properties_result:
        prop_id = prop['id']
        
        prop_units_sql = text("""
            SELECT 
                COUNT(DISTINCT u.id) as total_units,
                COUNT(DISTINCT CASE 
                    WHEN tu.id IS NOT NULL AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
                    THEN u.id 
                END) as occupied_units,
                COALESCE(SUM(CASE 
                    WHEN tu.id IS NOT NULL AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
                    THEN tu.monthly_rent 
                    ELSE 0 
                END), 0) as revenue
            FROM units u
            LEFT JOIN tenant_units tu ON tu.unit_id = u.id 
                AND (tu.move_out_date IS NULL OR tu.move_out_date > CURDATE())
            WHERE u.property_id = :prop_id
        """)
        prop_units_result = db.session.execute(prop_units_sql, {'prop_id': prop_id}).mappings().first()
        
        prop_total_units = int(prop_units_result['total_units']) if prop_units_result else 0
        prop_occupied_units = int(prop_units_result['occupied_units']) if prop_units_result else 0
        prop_revenue = float(prop_units_result['revenue']) if prop_units_result else 0.0
        prop_occupancy = round((prop_occupied_units / prop_total_units * 100), 2) if prop_total_units > 0 else 0.0
        
        property_performance.append({
            'property': prop.get('title') or prop.get('building_name') or f'Property {prop_id}',
            'occupancy': prop_occupancy,
            'revenue': prop_revenue
        })
    
    # Monthly data
    monthly_data = []
    for i in range(3):
        month_date = datetime.utcnow() - timedelta(days=30 * (2 - i))
        month_start = month_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if i == 2:
            month_end = datetime.utcnow()
        else:
            next_month = month_date + timedelta(days=32)
            month_end = next_month.replace(day=1) - timedelta(days=1)
        
        month_revenue_sql = text("""
            SELECT COALESCE(SUM(tu.monthly_rent), 0) as revenue
            FROM tenant_units tu
            INNER JOIN units u ON u.id = tu.unit_id
            WHERE u.property_id IN :property_ids
            AND tu.move_in_date <= :month_end
            AND (tu.move_out_date IS NULL OR tu.move_out_date >= :month_start)
        """)
        month_revenue_result = db.session.execute(
            month_revenue_sql,
            {'property_ids': property_ids_tuple, 'month_start': month_start, 'month_end': month_end}
        ).mappings().first()
        month_revenue = float(month_revenue_result['revenue']) if month_revenue_result else 0.0
        
        month_expenses_sql = text("""
            SELECT COALESCE(SUM(sb.amount), 0) as expenses
            FROM subscription_bills sb
            WHERE sb.user_id = :user_id
            AND sb.status IN ('paid', 'verified', 'completed')
            AND (
                (sb.billing_period_start >= :month_start AND sb.billing_period_start <= :month_end)
                OR (sb.payment_date >= :month_start AND sb.payment_date <= :month_end)
                OR (sb.created_at >= :month_start AND sb.created_at <= :month_end)
            )
        """)
        month_expenses_result = db.session.execute(
            month_expenses_sql,
            {'user_id': current_user.id, 'month_start': month_start, 'month_end': month_end}
        ).mappings().first()
        month_expenses = float(month_expenses_result['expenses']) if month_expenses_result else 0.0
        month_expenses = round(month_expenses, 2)
        
        monthly_data.append({
            'month': month_date.strftime('%B %Y'),
            'revenue': round(month_revenue, 2),
            'expenses': month_expenses
        })
    
    return {
        'properties': [dict(p) for p in properties_result],
        'totalRevenue': round(total_revenue, 2),
        'totalExpenses': round(total_expenses, 2),
        'netIncome': round(net_income, 2),
        'occupancyRate': occupancy_rate,
        'maintenanceRequests': maintenance_requests,
        'tenantSatisfaction': tenant_satisfaction,
        'monthlyData': monthly_data,
        'propertyPerformance': property_performance,
        'period': period,
        'property_filter': property_filter,
        'generated_at': datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')
    }


@manager_analytics_bp.route('/download/pdf', methods=['GET'])
@manager_required
def download_pdf_report(current_user):
    """Download analytics report as PDF."""
    if not REPORTLAB_AVAILABLE:
        return jsonify({'error': 'PDF generation not available. Please install reportlab.'}), 503
    
    try:
        property_filter = request.args.get('property', 'all')
        period = request.args.get('period', 'month')
        report_type = request.args.get('report_type', 'overview')
        
        data = _get_analytics_data(current_user, property_filter, period)
        
        # Create PDF in memory
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=1  # Center
        )
        story.append(Paragraph("Property Analytics Report", title_style))
        
        # Report info
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#666666'),
            alignment=1
        )
        period_label = period.replace('_', ' ').title()
        property_label = 'All Properties' if property_filter == 'all' else f'Property ID: {property_filter}'
        story.append(Paragraph(f"Period: {period_label} | {property_label}", info_style))
        story.append(Paragraph(f"Generated: {data['generated_at']}", info_style))
        story.append(Spacer(1, 0.3*inch))
        
        # Key Metrics Table
        metrics_data = [
            ['Metric', 'Value'],
            ['Total Revenue', f"₱{data['totalRevenue']:,.2f}"],
            ['Total Expenses', f"₱{data['totalExpenses']:,.2f}"],
            ['Net Income', f"₱{data['netIncome']:,.2f}"],
            ['Occupancy Rate', f"{data['occupancyRate']:.2f}%"],
            ['Maintenance Requests', str(data['maintenanceRequests'])],
            ['Tenant Satisfaction', f"{data['tenantSatisfaction']:.1f}/5.0"]
        ]
        
        metrics_table = Table(metrics_data, colWidths=[3*inch, 2*inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a1a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ]))
        story.append(Paragraph("Key Metrics", styles['Heading2']))
        story.append(Spacer(1, 0.1*inch))
        story.append(metrics_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Monthly Data
        if data['monthlyData']:
            story.append(Paragraph("Monthly Revenue & Expenses", styles['Heading2']))
            story.append(Spacer(1, 0.1*inch))
            monthly_data_table = [['Month', 'Revenue', 'Expenses', 'Net']]
            for month in data['monthlyData']:
                net = month['revenue'] - month['expenses']
                monthly_data_table.append([
                    month['month'],
                    f"₱{month['revenue']:,.2f}",
                    f"₱{month['expenses']:,.2f}",
                    f"₱{net:,.2f}"
                ])
            
            monthly_table = Table(monthly_data_table, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
            monthly_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a1a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
            ]))
            story.append(monthly_table)
            story.append(Spacer(1, 0.3*inch))
        
        # Property Performance
        if data['propertyPerformance']:
            story.append(Paragraph("Property Performance", styles['Heading2']))
            story.append(Spacer(1, 0.1*inch))
            prop_data_table = [['Property', 'Occupancy %', 'Monthly Revenue']]
            for prop in data['propertyPerformance']:
                prop_data_table.append([
                    prop['property'],
                    f"{prop['occupancy']:.2f}%",
                    f"₱{prop['revenue']:,.2f}"
                ])
            
            prop_table = Table(prop_data_table, colWidths=[3*inch, 1.5*inch, 1.5*inch])
            prop_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a1a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
            ]))
            story.append(prop_table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        # Generate filename
        period_str = period.replace(' ', '_')
        filename = f"analytics_report_{period_str}_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        current_app.logger.error(f'PDF report generation error: {e}', exc_info=True)
        return handle_api_error(500, 'Failed to generate PDF report')


@manager_analytics_bp.route('/download/excel', methods=['GET'])
@manager_required
def download_excel_report(current_user):
    """Download analytics report as Excel."""
    if not OPENPYXL_AVAILABLE:
        return jsonify({'error': 'Excel generation not available. Please install openpyxl.'}), 503
    
    try:
        property_filter = request.args.get('property', 'all')
        period = request.args.get('period', 'month')
        
        data = _get_analytics_data(current_user, property_filter, period)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Analytics Report"
        
        # Styles
        header_fill = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=16)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = "Property Analytics Report"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align
        
        # Report info
        period_label = period.replace('_', ' ').title()
        property_label = 'All Properties' if property_filter == 'all' else f'Property ID: {property_filter}'
        ws['A2'] = f"Period: {period_label} | {property_label}"
        ws['A3'] = f"Generated: {data['generated_at']}"
        
        row = 5
        
        # Key Metrics
        ws[f'A{row}'] = "Key Metrics"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        metrics_headers = ['Metric', 'Value']
        for col, header in enumerate(metrics_headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align
        
        row += 1
        metrics_data = [
            ['Total Revenue', f"₱{data['totalRevenue']:,.2f}"],
            ['Total Expenses', f"₱{data['totalExpenses']:,.2f}"],
            ['Net Income', f"₱{data['netIncome']:,.2f}"],
            ['Occupancy Rate', f"{data['occupancyRate']:.2f}%"],
            ['Maintenance Requests', str(data['maintenanceRequests'])],
            ['Tenant Satisfaction', f"{data['tenantSatisfaction']:.1f}/5.0"]
        ]
        
        for metric_row in metrics_data:
            for col, value in enumerate(metric_row, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = border
            row += 1
        
        row += 2
        
        # Monthly Data
        if data['monthlyData']:
            ws[f'A{row}'] = "Monthly Revenue & Expenses"
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 1
            
            monthly_headers = ['Month', 'Revenue', 'Expenses', 'Net']
            for col, header in enumerate(monthly_headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center_align
            
            row += 1
            for month in data['monthlyData']:
                net = month['revenue'] - month['expenses']
                ws.cell(row=row, column=1, value=month['month']).border = border
                ws.cell(row=row, column=2, value=f"₱{month['revenue']:,.2f}").border = border
                ws.cell(row=row, column=3, value=f"₱{month['expenses']:,.2f}").border = border
                ws.cell(row=row, column=4, value=f"₱{net:,.2f}").border = border
                row += 1
            
            row += 1
        
        # Property Performance
        if data['propertyPerformance']:
            ws[f'A{row}'] = "Property Performance"
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 1
            
            prop_headers = ['Property', 'Occupancy %', 'Monthly Revenue']
            for col, header in enumerate(prop_headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center_align
            
            row += 1
            for prop in data['propertyPerformance']:
                ws.cell(row=row, column=1, value=prop['property']).border = border
                ws.cell(row=row, column=2, value=f"{prop['occupancy']:.2f}%").border = border
                ws.cell(row=row, column=3, value=f"₱{prop['revenue']:,.2f}").border = border
                row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Generate filename
        period_str = period.replace(' ', '_')
        filename = f"analytics_report_{period_str}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        current_app.logger.error(f'Excel report generation error: {e}', exc_info=True)
        return handle_api_error(500, 'Failed to generate Excel report')


@manager_analytics_bp.route('/download/csv', methods=['GET'])
@manager_required
def download_csv_report(current_user):
    """Download analytics report as CSV."""
    try:
        property_filter = request.args.get('property', 'all')
        period = request.args.get('period', 'month')
        
        data = _get_analytics_data(current_user, property_filter, period)
        
        # Create CSV in memory
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        
        # Header
        writer.writerow(["Property Analytics Report"])
        period_label = period.replace('_', ' ').title()
        property_label = 'All Properties' if property_filter == 'all' else f'Property ID: {property_filter}'
        writer.writerow([f"Period: {period_label} | {property_label}"])
        writer.writerow([f"Generated: {data['generated_at']}"])
        writer.writerow([])
        
        # Key Metrics
        writer.writerow(["Key Metrics"])
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Total Revenue", f"₱{data['totalRevenue']:,.2f}"])
        writer.writerow(["Total Expenses", f"₱{data['totalExpenses']:,.2f}"])
        writer.writerow(["Net Income", f"₱{data['netIncome']:,.2f}"])
        writer.writerow(["Occupancy Rate", f"{data['occupancyRate']:.2f}%"])
        writer.writerow(["Maintenance Requests", str(data['maintenanceRequests'])])
        writer.writerow(["Tenant Satisfaction", f"{data['tenantSatisfaction']:.1f}/5.0"])
        writer.writerow([])
        
        # Monthly Data
        if data['monthlyData']:
            writer.writerow(["Monthly Revenue & Expenses"])
            writer.writerow(["Month", "Revenue", "Expenses", "Net"])
            for month in data['monthlyData']:
                net = month['revenue'] - month['expenses']
                writer.writerow([
                    month['month'],
                    f"₱{month['revenue']:,.2f}",
                    f"₱{month['expenses']:,.2f}",
                    f"₱{net:,.2f}"
                ])
            writer.writerow([])
        
        # Property Performance
        if data['propertyPerformance']:
            writer.writerow(["Property Performance"])
            writer.writerow(["Property", "Occupancy %", "Monthly Revenue"])
            for prop in data['propertyPerformance']:
                writer.writerow([
                    prop['property'],
                    f"{prop['occupancy']:.2f}%",
                    f"₱{prop['revenue']:,.2f}"
                ])
        
        # Convert to bytes
        csv_bytes = buffer.getvalue().encode('utf-8-sig')  # UTF-8 with BOM for Excel compatibility
        buffer.close()
        
        # Create response
        response = make_response(csv_bytes)
        period_str = period.replace(' ', '_')
        filename = f"analytics_report_{period_str}_{datetime.utcnow().strftime('%Y%m%d')}.csv"
        response.headers['Content-Type'] = 'text/csv; charset=utf-8-sig'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        current_app.logger.error(f'CSV report generation error: {e}', exc_info=True)
        return handle_api_error(500, 'Failed to generate CSV report')
